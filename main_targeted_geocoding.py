import os
import sys
import json
import time
import argparse
import logging
import urllib.request
import urllib.parse
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from tqdm import tqdm

# Import Phase 2 submodules
from src.geocoding_target_selector import identify_geocoding_targets
from src.address_enhancement import enhance_chemist_address_record, generate_geocoding_query
from src.pilot_sampler import draw_stratified_pilot_sample
from src.geocoding_precision import classify_precision_from_nominatim_raw, get_fallback_precision_by_query_type
from src.geocoding_acceptance import evaluate_geocoding_acceptance, calculate_haversine_single
from src.coordinate_promotion import calculate_proposed_fields, promote_proposed_coordinates
from src.pincode_reference import load_pincode_reference
from config import Config

# Setup logger
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger("TargetedGeocoding")

class TargetedGeocodingEngine:
    def __init__(self, mode="OFFLINE_ONLY", cache_filepath="reference_data/geocoding_cache.json"):
        self.mode = mode
        self.cache_filepath = cache_filepath
        self.cache = {}
        self.api_call_count = 0
        self.cache_hit_count = 0
        self._load_cache()

    def _load_cache(self):
        if os.path.exists(self.cache_filepath):
            try:
                with open(self.cache_filepath, 'r', encoding='utf-8') as f:
                    self.cache = json.load(f)
                logger.info(f"Loaded {len(self.cache)} entries from cache.")
            except Exception as e:
                logger.warning(f"Error loading cache: {e}. Starting fresh.")
                self.cache = {}
        else:
            dir_name = os.path.dirname(self.cache_filepath)
            if dir_name:
                os.makedirs(dir_name, exist_ok=True)
            self.cache = {}

    def save_cache(self):
        try:
            with open(self.cache_filepath, 'w', encoding='utf-8') as f:
                json.dump(self.cache, f, indent=2)
        except Exception as e:
            logger.error(f"Error saving cache: {e}")

    def geocode_query(self, query, query_type):
        """
        Executes lookup for a single query using cache first, then API if online.
        Returns: lat, lon, precision_class, confidence, is_cache_hit, status_message
        """
        query_clean = " ".join(query.split()).upper()
        
        # Check cache
        if query_clean in self.cache:
            self.cache_hit_count += 1
            cached = self.cache[query_clean]
            if cached.get('success', False):
                lat = cached['lat']
                lon = cached['lon']
                precision = cached.get('result_type', get_fallback_precision_by_query_type(cached.get('query_type', query_type)))
                conf = cached.get('confidence_score', 0.5)
                return lat, lon, precision, conf, True, "SUCCESS"
            else:
                return None, None, "FAILED", 0.0, True, f"Cached failure: {cached.get('error', 'Unknown error')}"

        # Cache miss - if OFFLINE_ONLY, do not execute API calls
        if self.mode == "OFFLINE_ONLY":
            return None, None, "FAILED", 0.0, False, "Geocoding is offline (cache miss)"

        # Check API sample limit (50 calls limit for public Nominatim safety)
        if self.mode == "PUBLIC_NOMINATIM_SAMPLE_ONLY" and self.api_call_count >= 50:
            return None, None, "FAILED", 0.0, False, "Public API call limit of 50 reached"

        # Execute API Lookup
        logger.info(f"Cache miss. Executing API call for: {query[:60]}...")
        data, success, error_msg = self._call_external_api(query)
        self.api_call_count += 1
        
        if success:
            lat = float(data['lat'])
            lon = float(data['lon'])
            precision = classify_precision_from_nominatim_raw(data)
            
            # Map default confidence
            conf = 0.95 if precision in ["SHOP_LEVEL", "BUILDING_LEVEL"] else (
                   0.80 if precision == "STREET_LEVEL" else (
                   0.60 if precision == "LOCALITY_LEVEL" else 0.30))
            
            self.cache[query_clean] = {
                'success': True,
                'lat': lat,
                'lon': lon,
                'result_type': precision,
                'class': data.get('class', ''),
                'type': data.get('type', ''),
                'address': data.get('address', {}),
                'confidence_score': conf,
                'query_type': query_type,
                'provider': self.mode,
                'timestamp': time.time()
            }
            self.save_cache()
            return lat, lon, precision, conf, False, "SUCCESS"
        else:
            self.cache[query_clean] = {
                'success': False,
                'error': error_msg,
                'timestamp': time.time()
            }
            self.save_cache()
            return None, None, "FAILED", 0.0, False, error_msg

    def _call_external_api(self, query):
        """
        Executes standard REST call to public OpenStreetMap Nominatim endpoint.
        Respects rate limits (minimum 1.2s delay).
        """
        time.sleep(1.2)  # Defensive rate-limiting
        
        provider_url = "https://nominatim.openstreetmap.org/search"
        params = {
            'q': query,
            'format': 'json',
            'limit': 1,
            'addressdetails': 1
        }
        url_parts = urllib.parse.urlencode(params)
        full_url = f"{provider_url}?{url_parts}"
        
        try:
            req = urllib.request.Request(
                full_url,
                headers={
                    'User-Agent': 'ChemistQualityPipeline/2.0 (targeted_cleansing_agent; contact: analyst@chemistmaster.local)',
                    'Accept-Language': 'en'
                }
            )
            with urllib.request.urlopen(req) as response:
                if response.status == 200:
                    data = json.loads(response.read().decode('utf-8'))
                    if data:
                        return data[0], True, ""
                    else:
                        return None, False, "No results found"
                else:
                    return None, False, f"HTTP Error {response.status}"
        except Exception as e:
            return None, False, str(e)

def calculate_city_medians(df):
    """
    Computes median coordinates of eligible records per city.
    """
    eligible = df[df['eligible_for_routing_flag'] == True]
    medians = eligible.groupby('normalized_city')[['final_lat', 'final_long']].median().to_dict('index')
    city_medians = {city: (coords['final_lat'], coords['final_long']) for city, coords in medians.items()}
    return city_medians

def write_formatted_excel(df, filepath, sheet_name):
    """
    Saves dataframe to Excel using openpyxl with formatted header, auto-fit, and filter filters.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Header styling
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    ws.append(list(df.columns))
    for col_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    # Append values
    clean_vals = df.replace({np.nan: None}).values.tolist()
    for row in clean_vals:
        ws.append(row)
        
    # Freeze header and enable auto filters
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(df.shape[1])}{df.shape[0] + 1}"
    
    # Auto-fit columns
    for col in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = min(40, max(max_len + 3, 10))
        
    wb.save(filepath)

def main():
    parser = argparse.ArgumentParser(description="Chemist Targeted Geocoding & Address Enhancement Pipeline")
    parser.add_argument("--input_file", type=str, default="outputs/chemist_master_cleaned_enriched.csv", help="Input Phase 1 cleaned file")
    parser.add_argument("--sample_mode", type=str, default="TRUE", help="Run in Sample/Pilot mode (TRUE or FALSE)")
    parser.add_argument("--sample_size", type=int, default=1000, help="Stratified sample size for pilot runs")
    parser.add_argument("--geocoding_mode", type=str, default="OFFLINE_ONLY", help="Geocoding mode (OFFLINE_ONLY, PUBLIC_NOMINATIM_SAMPLE_ONLY)")
    parser.add_argument("--apply_geocoding_updates", type=str, default="FALSE", help="Promote coordinates to final_lat/final_long (TRUE or FALSE)")
    parser.add_argument("--outputs_dir", type=str, default="outputs", help="Outputs target folder")
    
    args = parser.parse_args()
    
    start_time = time.time()
    run_timestamp = time.strftime('%Y-%m-%d %H:%M:%S')
    
    sample_mode_bool = args.sample_mode.upper() == "TRUE"
    apply_updates_bool = args.apply_geocoding_updates.upper() == "TRUE"
    
    os.makedirs(args.outputs_dir, exist_ok=True)
    
    logger.info("================================================================================")
    logger.info("CHEMIST TARGETED GEOCODING & ENHANCEMENT PIPELINE STARTING")
    logger.info("================================================================================")
    logger.info(f"Input Master:            {args.input_file}")
    logger.info(f"Pilot Sample Mode:       {sample_mode_bool}")
    if sample_mode_bool:
        logger.info(f"Sample Size:             {args.sample_size}")
    logger.info(f"Geocoding Mode:          {args.geocoding_mode}")
    logger.info(f"Apply Updates Promoted:  {apply_updates_bool}")
    logger.info(f"Outputs Folder:          {args.outputs_dir}")
    
    # 1. Load Input cleaned master
    if not os.path.exists(args.input_file):
        logger.critical(f"Input file '{args.input_file}' not found! Please execute Phase 1 pipeline first.")
        sys.exit(1)
        
    logger.info(f"Loading {args.input_file}...")
    df_master = pd.read_csv(args.input_file, dtype={'chemist_id': str, 'original_pincode': str, 'validated_pincode': str}, low_memory=False)
    
    # 2. Select targeted problematic records
    logger.info("Step 1: Identifying targeted records for re-geocoding audit...")
    df_targets = identify_geocoding_targets(df_master)
    logger.info(f"Found {len(df_targets)} chemist records requiring re-geocoding review.")
    
    # 3. Enhance address fields and build query queries
    logger.info("Step 2: Performing safe address correction expansions and query composition...")
    enhanced_addrs = []
    correction_statuses = []
    correction_notes = []
    
    for idx, row in df_targets.iterrows():
        enriched_addr, status, notes = enhance_chemist_address_record(
            row['original_address'], row['cleaned_address'], row['chemist_name_original'],
            row['normalized_city'], row['normalized_state'], row['validated_pincode']
        )
        enhanced_addrs.append(enriched_addr)
        correction_statuses.append(status)
        correction_notes.append(notes)
        
    df_targets['enriched_geocoding_address'] = enhanced_addrs
    df_targets['address_correction_status'] = correction_statuses
    df_targets['address_correction_notes'] = correction_notes
    
    # Export full target list prior to any geocoding attempts
    target_list_csv = os.path.join(args.outputs_dir, "chemist_regeocoding_target_list.csv")
    target_list_xlsx = os.path.join(args.outputs_dir, "chemist_regeocoding_target_list.xlsx")
    
    logger.info(f"Exporting raw target list to {target_list_csv}...")
    df_targets.to_csv(target_list_csv, index=False, encoding='utf-8')
    write_formatted_excel(df_targets, target_list_xlsx, "Re-geocoding Target List")
    
    # 4. Extract sample if Pilot mode is selected
    if sample_mode_bool:
        logger.info(f"Step 3: Compiling stratified pilot sample of size {args.sample_size}...")
        df_selected = draw_stratified_pilot_sample(df_targets, args.sample_size)
    else:
        logger.info("Step 3: Running full targeted geocoding on all target records...")
        df_selected = df_targets.copy()
        
    # 5. Load pincode ref database for distance checks
    logger.info("Loading reference pincode directory for validation checking...")
    pincode_ref = load_pincode_reference("reference_data/pincode_master.csv")
    
    # 6. Calculate city medians
    city_medians = calculate_city_medians(df_master)
    
    # 7. Core Geocoding Loop (Level-by-Level query strategy)
    logger.info("Step 4: Beginning query strategy lookup loops...")
    geocoder = TargetedGeocodingEngine(args.geocoding_mode)
    
    geocoding_attempted_flag = []
    geocoding_provider = []
    geocoding_query_used = []
    geocoding_query_level = []
    geocoding_status = []
    geocoding_result_lat = []
    geocoding_result_long = []
    geocoding_result_type = []
    geocoding_confidence_score = []
    geocoding_error_message = []
    geocoding_cache_hit_flag = []
    
    # Load config bounds
    config = Config()
    
    pbar = tqdm(total=len(df_selected), desc="Re-geocoding Targets")
    
    for idx, row in df_selected.iterrows():
        lat = np.nan
        lon = np.nan
        precision = "FAILED"
        conf = 0.0
        query_used = ""
        level_used = 0
        provider = "NONE"
        status = "FAILED"
        err_msg = ""
        is_hit = False
        
        # Check strict-to-loose Level 1-5
        for lvl in range(1, 6):
            query = generate_geocoding_query(
                lvl, row['chemist_name_original'], row['cleaned_address'],
                row['normalized_city'], row['normalized_state'], row['validated_pincode']
            )
            if not query:
                continue
                
            lvl_lat, lvl_lon, lvl_prec, lvl_conf, lvl_hit, lvl_status = geocoder.geocode_query(query, f"level_{lvl}")
            
            if lvl_status == "SUCCESS":
                lat = lvl_lat
                lon = lvl_lon
                precision = lvl_prec
                conf = lvl_conf
                query_used = query
                level_used = lvl
                provider = geocoder.mode
                status = "SUCCESS"
                is_hit = lvl_hit
                break
            else:
                err_msg = lvl_status  # Capture last failure message
                
        geocoding_attempted_flag.append(True)
        geocoding_provider.append(provider)
        geocoding_query_used.append(query_used)
        geocoding_query_level.append(level_used)
        geocoding_status.append(status)
        geocoding_result_lat.append(lat)
        geocoding_result_long.append(lon)
        geocoding_result_type.append(precision)
        geocoding_confidence_score.append(conf)
        geocoding_error_message.append(err_msg if status == "FAILED" else "")
        geocoding_cache_hit_flag.append(is_hit)
        
        pbar.update(1)
        
    pbar.close()
    
    df_selected['geocoding_attempted_flag'] = geocoding_attempted_flag
    df_selected['geocoding_provider'] = geocoding_provider
    df_selected['geocoding_query_used'] = geocoding_query_used
    df_selected['geocoding_query_level'] = geocoding_query_level
    df_selected['geocoding_status'] = geocoding_status
    df_selected['geocoding_result_lat'] = geocoding_result_lat
    df_selected['geocoding_result_long'] = geocoding_result_long
    df_selected['geocoding_result_type'] = geocoding_result_type
    df_selected['geocoding_confidence_score'] = geocoding_confidence_score
    df_selected['geocoding_error_message'] = geocoding_error_message
    df_selected['geocoding_cache_hit_flag'] = geocoding_cache_hit_flag
    
    # 8. Distance metrics
    logger.info("Step 5: Computing geographic distances to median center and original coords...")
    df_selected['distance_original_to_geocoded_km'] = df_selected.apply(
        lambda r: calculate_haversine_single(r['numeric_original_lat'], r['numeric_original_long'], r['geocoding_result_lat'], r['geocoding_result_long']), axis=1
    )
    
    df_selected['distance_geocoded_to_city_median_km'] = df_selected.apply(
        lambda r: calculate_haversine_single(r['geocoding_result_lat'], r['geocoding_result_long'], city_medians.get(r['normalized_city'], (np.nan, np.nan))[0], city_medians.get(r['normalized_city'], (np.nan, np.nan))[1]), axis=1
    )
    
    # 9. Validation & Acceptance Decisions
    logger.info("Step 6: Executing coordinate validation and acceptance rule checking...")
    decisions = []
    reasons = []
    
    for idx, row in df_selected.iterrows():
        dec, rsn = evaluate_geocoding_acceptance(
            row, row['geocoding_result_lat'], row['geocoding_result_long'],
            row['geocoding_result_type'], config, pincode_ref, city_medians
        )
        decisions.append(dec)
        reasons.append(rsn)
        
    df_selected['geocoding_acceptance_decision'] = decisions
    df_selected['geocoding_acceptance_reason'] = reasons
    
    # 10. Calculate Proposed Fields
    logger.info("Step 7: Compiling proposed quality score enhancements...")
    df_selected = calculate_proposed_fields(df_selected)
    
    # 11. Promote coordinates if updates are active
    df_selected = promote_proposed_coordinates(df_selected, apply_updates_bool)
    
    # If updates are applied, update the main df_master values too!
    if apply_updates_bool:
        logger.info("Applying geocoding updates into the main master copy...")
        # Index on record index
        df_master = df_master.set_index('chemist_record_index')
        df_selected_idx = df_selected.set_index('chemist_record_index')
        
        # Override columns in master with the selected rows
        cols_to_update = [
            'final_lat', 'final_long', 'coordinate_source', 'routing_reliability_bucket',
            'routing_risk_flag', 'overall_location_confidence_score'
        ]
        for col in cols_to_update:
            df_master.update(df_selected_idx[col])
            
        # Re-save master copy
        df_master = df_master.reset_index()
        df_master.to_csv(args.input_file, index=False, encoding='utf-8')
        logger.info("Updated chemist_master_cleaned_enriched.csv successfully.")
        
    # 12. Write Outputs
    logger.info("Step 8: Exporting candidate tables and decision logs...")
    
    # Determine result files depending on run mode
    res_prefix = "geocoding_pilot" if sample_mode_bool else "chemist_geocoding_candidate"
    
    cand_csv = os.path.join(args.outputs_dir, f"{res_prefix}_results.csv")
    cand_xlsx = os.path.join(args.outputs_dir, f"{res_prefix}_results.xlsx")
    df_selected.to_csv(cand_csv, index=False, encoding='utf-8')
    write_formatted_excel(df_selected, cand_xlsx, "Geocoding Candidates")
    
    # Save cache CSV
    cache_csv = os.path.join(args.outputs_dir, "geocoding_cache.csv")
    cache_rows = []
    for q_clean, val in geocoder.cache.items():
        if val.get('success', False):
            cache_rows.append({
                'Cleaned_Query': q_clean,
                'lat': val.get('lat'),
                'lon': val.get('lon'),
                'precision': val.get('result_type', ''),
                'query_type': val.get('query_type', ''),
                'timestamp': val.get('timestamp')
            })
    pd.DataFrame(cache_rows).to_csv(cache_csv, index=False, encoding='utf-8')
    
    # Save configurations
    config_path = os.path.join(args.outputs_dir, "geocoding_config_used.json")
    with open(config_path, 'w', encoding='utf-8') as f:
        json.dump(vars(args), f, indent=2)
        
    # Coordinate Proposals file (ACCEPT_REPLACE_FINAL_COORDS, ACCEPT_WITH_CAUTION, PINCODE_CENTROID_ONLY, KEEP_ORIGINAL_WITH_FLAGS)
    proposal_mask = df_selected['geocoding_acceptance_decision'].isin([
        "ACCEPT_REPLACE_FINAL_COORDS", "ACCEPT_WITH_CAUTION", "PINCODE_CENTROID_ONLY", "KEEP_ORIGINAL_WITH_FLAGS"
    ])
    df_prop = df_selected[proposal_mask].copy()
    
    prop_cols = [
        'chemist_record_index', 'chemist_id', 'chemist_name_original', 'original_address', 'cleaned_address',
        'enriched_geocoding_address', 'normalized_city', 'normalized_state', 'validated_pincode',
        'original_lat', 'original_long', 'previous_final_lat', 'previous_final_long',
        'geocoding_result_lat', 'geocoding_result_long', 'proposed_final_lat', 'proposed_final_long',
        'coordinate_source', 'proposed_coordinate_source', 'distance_original_to_geocoded_km',
        'geocoding_result_type', 'geocoding_confidence_score', 'geocoding_acceptance_decision',
        'geocoding_acceptance_reason', 'quality_bucket', 'routing_reliability_bucket',
        'proposed_routing_reliability_bucket', 'address_issue_flags', 'pincode_issue_flags', 'coordinate_issue_flags'
    ]
    prop_cols = [c for c in prop_cols if c in df_prop.columns]
    df_prop_sub = df_prop[prop_cols]
    
    prop_csv = os.path.join(args.outputs_dir, "chemist_coordinate_update_proposals.csv")
    prop_xlsx = os.path.join(args.outputs_dir, "chemist_coordinate_update_proposals.xlsx")
    df_prop_sub.to_csv(prop_csv, index=False, encoding='utf-8')
    write_formatted_excel(df_prop_sub, prop_xlsx, "Coordinate Proposals")
    
    # Manual Review Queue (MANUAL_REVIEW_REQUIRED, REJECT_LOW_CONFIDENCE, GEOCODING_FAILED, and conflict boundaries)
    mr_mask = df_selected['geocoding_acceptance_decision'].isin([
        "MANUAL_REVIEW_REQUIRED", "REJECT_LOW_CONFIDENCE", "GEOCODING_FAILED"
    ]) | (
        # High distance difference but both might be flagged
        (df_selected['distance_original_to_geocoded_km'] > 10.0) & 
        (df_selected['geocoding_acceptance_decision'].isin(["ACCEPT_REPLACE_FINAL_COORDS", "ACCEPT_WITH_CAUTION"]))
    )
    df_mr = df_selected[mr_mask].copy()
    
    # Add review columns
    df_mr['manual_review_reason'] = df_mr['geocoding_acceptance_reason']
    
    mr_pri = []
    mr_step = []
    for idx, row in df_mr.iterrows():
        dec = row['geocoding_acceptance_decision']
        if dec == "MANUAL_REVIEW_REQUIRED":
            mr_pri.append("P1_CRITICAL")
            mr_step.append("Verify specific landmarks on satellite maps manually")
        elif dec == "REJECT_LOW_CONFIDENCE":
            mr_pri.append("P2_HIGH")
            mr_step.append("Gather local medical sales representative address feedback")
        else:
            mr_pri.append("P3_MEDIUM")
            mr_step.append("Verify location spellings or look for synonymous shops")
            
    df_mr['manual_review_priority'] = mr_pri
    df_mr['suggested_next_step'] = mr_step
    
    mr_cols = [
        'chemist_record_index', 'chemist_id', 'chemist_name_original', 'cleaned_address',
        'normalized_city', 'normalized_state', 'validated_pincode',
        'original_lat', 'original_long', 'geocoding_result_lat', 'geocoding_result_long',
        'distance_original_to_geocoded_km', 'geocoding_result_type', 'geocoding_acceptance_decision',
        'manual_review_reason', 'manual_review_priority', 'suggested_next_step'
    ]
    mr_cols = [c for c in mr_cols if c in df_mr.columns]
    df_mr_sub = df_mr[mr_cols]
    
    mr_xlsx = os.path.join(args.outputs_dir, "chemist_geocoding_manual_review_queue.xlsx")
    write_formatted_excel(df_mr_sub, mr_xlsx, "Geocoding Manual Review")
    
    # Save decision counts breakdown
    acceptance_summary_csv = os.path.join(args.outputs_dir, "chemist_geocoding_acceptance_summary.csv")
    acceptance_summary_xlsx = os.path.join(args.outputs_dir, "chemist_geocoding_acceptance_summary.xlsx")
    dec_counts = df_selected['geocoding_acceptance_decision'].value_counts().reset_index()
    dec_counts.columns = ['Acceptance_Decision', 'Record_Count']
    dec_counts['Percentage'] = (dec_counts['Record_Count'] / len(df_selected) * 100).round(2)
    dec_counts.to_csv(acceptance_summary_csv, index=False, encoding='utf-8')
    write_formatted_excel(dec_counts, acceptance_summary_xlsx, "Acceptance Decisions")
    
    # 13. Write Text Summary report
    elapsed_time = time.time() - start_time
    sum_prefix = "geocoding_pilot" if sample_mode_bool else "geocoding_run"
    sum_filepath = os.path.join(args.outputs_dir, f"{sum_prefix}_summary.txt")
    
    total_attempted = len(df_selected)
    successful = (df_selected['geocoding_status'] == "SUCCESS").sum()
    failed = (df_selected['geocoding_status'] == "FAILED").sum()
    
    # Result precision counts
    prec_counts = df_selected['geocoding_result_type'].value_counts()
    prec_str = "\n".join([f"  - {k}: {v} ({v/total_attempted*100:.2f}%)" for k, v in prec_counts.items()])
    
    # Decision counts
    dec_counts_map = df_selected['geocoding_acceptance_decision'].value_counts()
    dec_str = "\n".join([f"  - {k}: {v} ({v/total_attempted*100:.2f}%)" for k, v in dec_counts_map.items()])
    
    avg_conf = df_selected[df_selected['geocoding_status'] == "SUCCESS"]['geocoding_confidence_score'].mean()
    avg_conf_val = f"{avg_conf:.2f}" if not pd.isna(avg_conf) else "N/A"
    
    summary_report = f"""================================================================================
CHEMIST TARGETED GEOCODING RUN SUMMARY ({'PILOT MODE' if sample_mode_bool else 'FULL TARGETED RUN'})
================================================================================
Run Timestamp:             {run_timestamp}
Geocoding Provider Mode:   {args.geocoding_mode}
Target Records Attempted:  {total_attempted}
Successful Geocodes:       {successful}
Failed Geocodes:           {failed}
Cache Hits:                {geocoder.cache_hit_count}
API calls made:            {geocoder.api_call_count}
Average Confidence Score:  {avg_conf_val}
Pipeline Runtime:          {elapsed_time:.2f} seconds

RESULT PRECISION TYPE DISTRIBUTION:
--------------------------------------------------------------------------------
{prec_str}

ACCEPTANCE DECISION DISTRIBUTION:
--------------------------------------------------------------------------------
{dec_str}

DECISION KEY METRIC SPLITS:
--------------------------------------------------------------------------------
- ACCEPT_REPLACE_FINAL_COORDS:   {dec_counts_map.get('ACCEPT_REPLACE_FINAL_COORDS', 0)}
- ACCEPT_WITH_CAUTION:          {dec_counts_map.get('ACCEPT_WITH_CAUTION', 0)}
- PINCODE_CENTROID_ONLY:         {dec_counts_map.get('PINCODE_CENTROID_ONLY', 0)}
- KEEP_ORIGINAL_WITH_FLAGS:      {dec_counts_map.get('KEEP_ORIGINAL_WITH_FLAGS', 0)}
- REJECT_LOW_CONFIDENCE:         {dec_counts_map.get('REJECT_LOW_CONFIDENCE', 0)}
- MANUAL_REVIEW_REQUIRED:        {dec_counts_map.get('MANUAL_REVIEW_REQUIRED', 0)}
- GEOCODING_FAILED:              {dec_counts_map.get('GEOCODING_FAILED', 0)}

WARNINGS & PROVIDER USAGE NOTES:
--------------------------------------------------------------------------------
1. In OFFLINE_ONLY mode, API lookup calls are suppressed. Missing items stay blank.
2. In PUBLIC_NOMINATIM_SAMPLE_ONLY mode, lookups are rate-limited to 1 request per 1.2s to protect IP. Maximum API call limit is capped at 50 per execution.
3. Coordinates are updated only when apply_geocoding_updates=TRUE. Otherwise, values remain strictly in proposed candidate columns to preserve auditability.
================================================================================
"""
    with open(sum_filepath, 'w', encoding='utf-8') as f:
        f.write(summary_report)
        
    logger.info(f"Targeted geocoding script complete! Execution time: {elapsed_time:.2f} seconds.")

if __name__ == "__main__":
    main()
